# ================================================
# 요구사항
# ================================================
# Python을 이용한 공공기관 메뉴 URL 및 디바이스별 캡처

# cd /SYSTEMS/PROG/PYTHON/IDE/src/talentPlatform/unitSys
# /SYSTEMS/LIB/anaconda3/envs/py39/bin/python /SYSTEMS/PROG/PYTHON/IDE/src/talentPlatform/unitSys/TalentPlatform-INDI2026-DaemonFramework-captWebPage.py
# nohup /SYSTEMS/LIB/anaconda3/envs/py39/bin/python /SYSTEMS/PROG/PYTHON/IDE/src/talentPlatform/unitSys/TalentPlatform-INDI2026-DaemonFramework-captWebPage.py &
# tail -f nohup.out

import argparse
import glob
import logging
import logging.handlers
import os
import platform
import sys
import traceback
import warnings
from datetime import datetime
import time
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import csv
import os
import pandas as pd
import re
import json
from datetime import datetime, timedelta
# from konlpy.tag import Okt
from collections import Counter
import pytz
import os
import sys
import urllib.request
import os
import sys
import requests
import json
from konlpy.tag import Okt
from newspaper import Article
import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
from google.cloud import bigquery
from google.oauth2 import service_account
import requests
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright
import time
from playwright.sync_api import sync_playwright
import time
from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
from bs4 import BeautifulSoup
import pandas as pd
import requests
from bs4 import BeautifulSoup
import pandas as pd
from urllib.parse import urljoin

# =================================================
# 사용자 매뉴얼
# =================================================
# [소스 코드의 실행 순서]
# 1. 초기 설정 : 폰트 설정
# 2. 유틸리티 함수 : 초기화 함수 (로그 설정, 초기 변수, 초기 전달인자 설정) 또는 자주 사용하는 함수
# 3. 주 프로그램 :부 프로그램을 호출
# 4. 부 프로그램 : 자료 처리를 위한 클래스로서 내부 함수 (초기 변수, 비즈니스 로직, 수행 프로그램 설정)
# 4.1. 환경 변수 설정 (로그 설정) : 로그 기록을 위한 설정 정보 읽기
# 4.2. 환경 변수 설정 (초기 변수) : 입력 경로 (inpPath) 및 출력 경로 (outPath) 등을 설정
# 4.3. 초기 변수 (Argument, Option) 설정 : 파이썬 실행 시 전달인자 설정 (pyhton3 *.py argv1 argv2 argv3 ...)
# 4.4. 비즈니스 로직 수행 : 단위 시스템 (unit 파일명)으로 관리 또는 비즈니스 로직 구현

# =================================================
# 1. 초기 설정
# =================================================
warnings.filterwarnings("ignore")

plt.rc('font', family='Malgun Gothic')
plt.rc('axes', unicode_minus=False)
# sns.set(font="Malgun Gothic", rc={"axes.unicode_minus": False}, style='darkgrid')

# 그래프에서 마이너스 글꼴 깨지는 문제에 대한 대처
mpl.rcParams['axes.unicode_minus'] = False

# 타임존 설정
tzKst = pytz.timezone('Asia/Seoul')
tzUtc = pytz.timezone('UTC')

# =================================================
# 2. 유틸리티 함수
# =================================================
# 로그 설정
def initLog(env=None, contextPath=None, prjName=None):
    if env is None: env = 'local'
    if contextPath is None: contextPath = os.getcwd()
    if prjName is None: prjName = 'test'

    saveLogFile = "{}/{}_{}_{}_{}_{}.log".format(
        contextPath if env in 'local' else os.path.join(contextPath, 'resources', 'log', prjName)
        , platform.system()
        , platform.machine()
        , platform.architecture()[0]
        , platform.node()
        , prjName
    )

    os.makedirs(os.path.dirname(saveLogFile), exist_ok=True)

    log = logging.getLogger(prjName)

    if len(log.handlers) > 0:
        return log

    format = logging.Formatter('%(asctime)s [%(name)s | %(lineno)d | %(filename)s] [%(levelname)-5.5s] %(message)s')

    streamHandler = logging.StreamHandler()
    fileHandler = logging.handlers.TimedRotatingFileHandler(filename=saveLogFile, when='midnight', interval=1, backupCount=30, encoding='utf-8')

    streamHandler.setFormatter(format)
    fileHandler.setFormatter(format)

    log.addHandler(streamHandler)
    log.addHandler(fileHandler)

    log.setLevel(level=logging.INFO)

    return log

#  초기 변수 설정
def initGlobalVar(env=None, contextPath=None, prjName=None):
    if env is None: env = 'local'
    if contextPath is None: contextPath = os.getcwd()
    if prjName is None: prjName = 'test'

    # 환경 변수 (local, 그 외)에 따라 전역 변수 (입력 자료, 출력 자료 등)를 동적으로 설정
    # 즉 local의 경우 현재 작업 경로 (contextPath)를 기준으로 설정
    # 그 외의 경우 contextPath/resources/input/prjName와 같은 동적으로 구성
    globalVar = {
        'prjName': prjName
        , 'sysOs': platform.system()
        , 'contextPath': contextPath
        , 'resPath': contextPath if env in 'local' else os.path.join(contextPath, 'resources')
        , 'cfgPath': contextPath if env in 'local' else os.path.join(contextPath, 'resources', 'config')
        , 'inpPath': contextPath if env in 'local' else os.path.join(contextPath, 'resources', 'input', prjName)
        , 'figPath': contextPath if env in 'local' else os.path.join(contextPath, 'resources', 'fig', prjName)
        , 'outPath': contextPath if env in 'local' else os.path.join(contextPath, 'resources', 'output', prjName)
        , 'movPath': contextPath if env in 'local' else os.path.join(contextPath, 'resources', 'movie', prjName)
        , 'logPath': contextPath if env in 'local' else os.path.join(contextPath, 'resources', 'log', prjName)
        , 'mapPath': contextPath if env in 'local' else os.path.join(contextPath, 'resources', 'mapInfo')
        , 'sysCfg': contextPath if env in 'local' else os.path.join(contextPath, 'resources', 'config', 'system.json')
        , 'seleniumPath': contextPath if env in 'local' else os.path.join(contextPath, 'resources', 'config', 'selenium')
        , 'fontPath': contextPath if env in 'local' else os.path.join(contextPath, 'resources', 'config', 'fontInfo')
    }

    return globalVar

#  초기 전달인자 설정
def initArgument(globalVar):
    parser = argparse.ArgumentParser()

    for i, argv in enumerate(sys.argv[1:]):
        if not argv.__contains__('--'): continue
        parser.add_argument(argv)

    inParInfo = vars(parser.parse_args())
    log.info(f"[CHECK] inParInfo : {inParInfo}")

    # 전역 변수에 할당
    for key, val in inParInfo.items():
        if val is None: continue
        if env not in 'local' and key.__contains__('Path'):
            os.makedirs(val, exist_ok=True)
        globalVar[key] = val

    return globalVar


# ================================================
# 4. 부 프로그램
# ================================================
class DtaProcess(object):

    # ================================================================================================
    # 환경변수 설정
    # ================================================================================================
    global env, contextPath, prjName, serviceName, log, globalVar

    # env = 'local'  # 로컬 : 원도우 환경, 작업환경 (현재 소스 코드 환경 시 .) 설정
    env = 'dev'  # 개발 : 원도우 환경, 작업환경 (사용자 환경 시 contextPath) 설정
    # env = 'oper'  # 운영 : 리눅스 환경, 작업환경 (사용자 환경 시 contextPath) 설정

    if platform.system() == 'Windows':
        contextPath = os.getcwd() if env in 'local' else 'E:/04. TalentPlatform/Github/TalentPlatform-Python'
    else:
        contextPath = os.getcwd() if env in 'local' else '/SYSTEMS/PROG/PYTHON/IDE'

    prjName = 'captWebPage'
    serviceName = 'INDI2026'

    # 4.1. 환경 변수 설정 (로그 설정)
    log = initLog(env, contextPath, prjName)

    # 4.2. 환경 변수 설정 (초기 변수)
    globalVar = initGlobalVar(env, contextPath, prjName)

    # ================================================================================================
    # 4.3. 초기 변수 (Argument, Option) 설정
    # ================================================================================================
    def __init__(self):

        log.info('[START] {}'.format("init"))

        try:
            # 초기 전달인자 설정 (파이썬 실행 시)
            # pyhton3 *.py argv1 argv2 argv3 ...
            initArgument(globalVar)

        except Exception as e:
            log.error(f"Exception : {str(e)}")
            raise e
        finally:
            log.info('[END] {}'.format("init"))

    # ================================================================================================
    # 4.4. 비즈니스 로직 수행
    # ================================================================================================
    def exec(self):

        log.info('[START] {}'.format("exec"))

        try:

            if (platform.system() == 'Windows'):
                pass
            else:
                globalVar['inpPath'] = '/DATA/INPUT'
                globalVar['outPath'] = '/DATA/OUTPUT'
                globalVar['figPath'] = '/DATA/FIG'

            # 옵션 설정
            sysOpt = {
                'headers': {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
                },
                'url': 'https://nmsc.kma.go.kr/homepage/html/main/main.do',
                'saveFile': '/DATA/OUTPUT/INDI2026/capture_with_images.xlsx',
            }

            # ==========================================================================================================
            # 공공기관 메뉴 URL 수집
            # ==========================================================================================================
            target_url = "https://nmsc.kma.go.kr/homepage/html/main/main.do"
            base_domain = "https://nmsc.kma.go.kr"
            log.info(f"Playwright 브라우저로 웹사이트 접속 중... ({target_url})")

            html_content = ""
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                page = browser.new_page()

                try:
                    page.goto(target_url, timeout=60000)

                    page.wait_for_selector('#menutop', timeout=10000)

                    page.wait_for_timeout(1000)

                    html_content = page.content()
                    log.info("브라우저 렌더링 완료 및 HTML 소스 추출 성공!")

                except Exception as e:
                    log.info(f"웹페이지 로딩 중 오류 발생: {e}")

                finally:
                    browser.close()

            soup = BeautifulSoup(html_content, 'html.parser')
            menu_top = soup.find('ul', id='menutop')

            if not menu_top:
                log.info("페이지에서 'menutop' 아이디를 찾을 수 없습니다.")

            parsed_data = []

            depth1_lis = menu_top.find_all('li', class_=lambda c: c and c.startswith('depth1-menu'), recursive=False)

            for d1 in depth1_lis:
                d1_a = d1.find('a', recursive=False)
                d1_name = d1_a.text.strip() if d1_a else ""

                depth2_ul = d1.find('ul', class_='depth2-list')
                if depth2_ul:
                    depth2_lis = depth2_ul.find_all('li', class_=lambda c: c and c.startswith('depth2-menu'),
                                                    recursive=False)

                    for d2 in depth2_lis:
                        d2_a = d2.find('a', recursive=False) or d2.find('a')
                        d2_name = d2_a.contents[0].strip() if isinstance(d2_a.contents[0], str) else d2_a.text.strip()

                        depth3_ul = d2.find('ul', class_='depth3-list')
                        if depth3_ul:
                            depth3_lis = depth3_ul.find_all('li', class_=lambda c: c and c.startswith('depth3-menu'))

                            for d3 in depth3_lis:
                                d3_a = d3.find('a', recursive=False) or d3.find('a')
                                nested_a = d3.find('p').find('a') if d3.find('p') else None

                                d3_name = d3_a.contents[0].strip() if isinstance(d3_a.contents[0],str) else d3_a.text.strip()
                                d3_name = d3_name.replace('(새창)', '').replace('새창', '').strip()

                                raw_link = nested_a.get('href') if nested_a else d3_a.get('href')
                                final_link = urljoin(base_domain, raw_link) if raw_link and raw_link != "void" else "링크 없음"

                                parsed_data.append([d1_name, d2_name, d3_name, f"{d1_name} > {d2_name} > {d3_name}", final_link])
                        else:
                            raw_link = d2_a.get('href')
                            final_link = urljoin(base_domain, raw_link) if raw_link and raw_link != "void" else "링크 없음"
                            parsed_data.append([d1_name, d2_name, "", f"{d1_name} > {d2_name}", final_link])
                else:
                    raw_link = d1_a.get('href')
                    final_link = urljoin(base_domain, raw_link) if raw_link and raw_link != "void" else "링크 없음"
                    parsed_data.append([d1_name, "", "", d1_name, final_link])

            columns = ["1뎁스", "2뎁스", "3뎁스(메뉴명)", "전체 메뉴경로", "URL 주소"]
            df = pd.DataFrame(parsed_data, columns=columns)

            condition = (df['전체 메뉴경로'] == '자료조회 > 위성 영상 > 국민생활 안전 및 편의')
            df.loc[condition, 'URL 주소'] = 'https://nmsc.kma.go.kr/exposition/exposition.html'

            # ==========================================================================================================
            # 디바이스별 캡처
            # ==========================================================================================================
            viewports = {
                "PC": {"width": 1920, "height": 1080},
                "Tablet_Landscape": {"width": 1024, "height": 768},
                "Tablet_Portrait": {"width": 768, "height": 1024},
                "Mobile": {"width": 375, "height": 812}
            }

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "디바이스별 전체화면 캡처"
            ws.views.sheetView[0].showGridLines = True

            header_font = Font(name="Malgun Gothic", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'),
                                 top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0'))

            headers = ["메뉴 경로", "URL", "PC", "태블릿 가로", "태블릿 세로", "모바일"]
            ws.append(headers)
            ws.row_dimensions[1].height = 30

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            log.info("DataFrame 기준 일괄 캡처 및 엑셀 삽입 작업을 시작합니다...")

            col_mapping = {"PC": "C", "Tablet_Landscape": "D", "Tablet_Portrait": "E", "Mobile": "F"}

            max_excel_row_height = 405
            target_img_height_px = int(max_excel_row_height / 0.75)
            row_idx = 2

            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)

                for index, row in df.iterrows():
                    menu_path = row['전체 메뉴경로']
                    url = row['URL 주소']

                    if pd.isna(url) or url == "링크 없음":
                        continue

                    log.info(f"[{index + 1}/{len(df)}] 캡처 진행 중: {menu_path}")

                    # A, B열 텍스트 데이터 입력
                    ws.cell(row=row_idx, column=1, value=menu_path).alignment = center_align
                    ws.cell(row=row_idx, column=1).border = thin_border
                    ws.cell(row=row_idx, column=2, value=url).alignment = center_align
                    ws.cell(row=row_idx, column=2).border = thin_border

                    for mode, size in viewports.items():
                        is_mobile_mode = ("Mobile" in mode or "Tablet" in mode)
                        context = browser.new_context(viewport=size, is_mobile=is_mobile_mode)
                        page = context.new_page()

                        file_name = f"capture_{index}_{mode}.png"

                        try:
                            page.goto(url, timeout=60000)
                            page.wait_for_timeout(3000)

                            page.screenshot(path=file_name, full_page=True)

                            if os.path.exists(file_name):
                                col_letter = col_mapping[mode]
                                img = ExcelImage(file_name)

                                # 비율 유지 축소 계산
                                original_width = img.width
                                original_height = img.height
                                ratio = target_img_height_px / original_height

                                img.height = target_img_height_px
                                img.width = int(original_width * ratio)

                                ws.add_image(img, f"{col_letter}{row_idx}")

                                current_col_width = ws.column_dimensions[col_letter].width or 10
                                new_col_width = (img.width / 7.5) + 3
                                if new_col_width > current_col_width:
                                    ws.column_dimensions[col_letter].width = new_col_width

                                ws.cell(row=row_idx, column=ord(col_letter) - 64).border = thin_border
                        except Exception as e:
                            log.info(f"[{mode}] 캡처 오류: {e}")
                        finally:
                            context.close()

                    ws.row_dimensions[row_idx].height = max_excel_row_height
                    row_idx += 1

                browser.close()

            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 50

            # ==========================================================================================================
            # 엑셀 파일 저장
            # ==========================================================================================================
            saveFile = sysOpt['saveFile']
            os.makedirs(os.path.dirname(saveFile), exist_ok=True)
            wb.save(saveFile)
            log.info(f'[CHECK] saveFile : {saveFile}')

        except Exception as e:
            log.error(f"Exception : {e}")
            raise e
        finally:
            log.info('[END] {}'.format("exec"))

# ================================================
# 3. 주 프로그램
# ================================================
if __name__ == '__main__':

    print('[START] {}'.format("main"))

    try:
        # 부 프로그램 호출
        subDtaProcess = DtaProcess()
        subDtaProcess.exec()

    except Exception as e:
        print(traceback.format_exc())
        sys.exit(1)

    finally:
        print('[END] {}'.format("main"))