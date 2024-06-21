# -*- coding: utf-8 -*-
import argparse
import glob
import logging
import logging.handlers
import os
import platform
import sys
import time
import traceback
import warnings
from datetime import datetime

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import pandas
import pandas as pd
import xarray as xr
import re

import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import ScalarFormatter
import seaborn as sns
from multiprocessing import Pool
import multiprocessing as mp

from openpyxl import load_workbook
from openpyxl.styles import Font
from copy import copy

# =================================================
# 사용자 매뉴얼
# =================================================
# [소스 코드의 실행 순서]
# 1. 초기 설정 : 폰트 설정
# 2. 유틸리티 함수 : 초기화 함수 (로그 설정, 초기 변수, 초기 전달인자 설정) 또는 자주 사용하는 함수
# 3. 주 프로그램 :부 프로그램을 호출
# 4. 부 프로그램 : 자료 처리를 위한 클래스로서 내부 함수 (초기 변수, 비즈니스 로직, 수행 프로그램 설정)
# 4.1. 환경 변수 설정 (로그 설정) : 로그 기록을 위한 설정 정보 읽기
# 4.2. 환경 변수 설정 (초기 변수) : 입력 경로 (inpPath) 및 출력 경로 (outPath) 등을 설정
# 4.3. 초기 변수 (Argument, Option) 설정 : 파이썬 실행 시 전달인자 설정 (pyhton3 *.py argv1 argv2 argv3 ...)
# 4.4. 비즈니스 로직 수행 : 단위 시스템 (unit 파일명)으로 관리 또는 비즈니스 로직 구현

# =================================================
# 1. 초기 설정
# =================================================
warnings.filterwarnings("ignore")

plt.rc('font', family='Malgun Gothic')
plt.rc('axes', unicode_minus=False)
# sns.set(font="Malgun Gothic", rc={"axes.unicode_minus": False}, style='darkgrid')

# 그래프에서 마이너스 글꼴 깨지는 문제에 대한 대처
mpl.rcParams['axes.unicode_minus'] = False

# =================================================
# 2. 유틸리티 함수
# =================================================
# 로그 설정
def initLog(env=None, contextPath=None, prjName=None):
    if env is None: env = 'local'
    if contextPath is None: contextPath = os.getcwd()
    if prjName is None: prjName = 'test'

    saveLogFile = "{}/{}_{}_{}_{}_{}_{}.log".format(
        contextPath if env in 'local' else os.path.join(contextPath, 'resources', 'log', prjName)
        , platform.system()
        , platform.machine()
        , platform.architecture()[0]
        , platform.node()
        , prjName
        , datetime.now().strftime("%Y%m%d")
    )

    os.makedirs(os.path.dirname(saveLogFile), exist_ok=True)

    # logger instance 생성
    log = logging.getLogger(prjName)

    if len(log.handlers) > 0:
        return log

    # format 생성
    format = logging.Formatter('%(asctime)s [%(name)s | %(lineno)d | %(filename)s] [%(levelname)-5.5s] %(message)s')

    # handler 생성
    streamHandler = logging.StreamHandler()
    fileHandler = logging.FileHandler(saveLogFile)

    # logger instance에 format 설정
    streamHandler.setFormatter(format)
    fileHandler.setFormatter(format)

    # logger instance에 handler 설정
    log.addHandler(streamHandler)
    log.addHandler(fileHandler)

    # logger instance로 log 기록
    log.setLevel(level=logging.INFO)

    return log


#  초기 변수 설정
def initGlobalVar(env=None, contextPath=None, prjName=None):
    if env is None: env = 'local'
    if contextPath is None: contextPath = os.getcwd()
    if prjName is None: prjName = 'test'

    # 환경 변수 (local, 그 외)에 따라 전역 변수 (입력 자료, 출력 자료 등)를 동적으로 설정
    # 즉 local의 경우 현재 작업 경로 (contextPath)를 기준으로 설정
    # 그 외의 경우 contextPath/resources/input/prjName와 같은 동적으로 구성
    globalVar = {
        'prjName': prjName
        , 'sysOs': platform.system()
        , 'contextPath': contextPath
        , 'resPath': contextPath if env in 'local' else os.path.join(contextPath, 'resources')
        , 'cfgPath': contextPath if env in 'local' else os.path.join(contextPath, 'resources', 'config')
        , 'inpPath': contextPath if env in 'local' else os.path.join(contextPath, 'resources', 'input', prjName)
        , 'figPath': contextPath if env in 'local' else os.path.join(contextPath, 'resources', 'fig', prjName)
        , 'outPath': contextPath if env in 'local' else os.path.join(contextPath, 'resources', 'output', prjName)
        , 'movPath': contextPath if env in 'local' else os.path.join(contextPath, 'resources', 'movie', prjName)
        , 'logPath': contextPath if env in 'local' else os.path.join(contextPath, 'resources', 'log', prjName)
        , 'mapPath': contextPath if env in 'local' else os.path.join(contextPath, 'resources', 'config', 'mapInfo')
        , 'sysPath': contextPath if env in 'local' else os.path.join(contextPath, 'resources', 'config', 'system.cfg')
        , 'seleniumPath': contextPath if env in 'local' else os.path.join(contextPath, 'resources', 'config', 'selenium')
        , 'fontPath': contextPath if env in 'local' else os.path.join(contextPath, 'resources', 'config', 'fontInfo')
    }

    return globalVar


#  초기 전달인자 설정
def initArgument(globalVar, inParams):
    # 원도우 또는 맥 환경
    if globalVar['sysOs'] in 'Windows' or globalVar['sysOs'] in 'Darwin':
        inParInfo = inParams

    # 리눅스 환경
    if globalVar['sysOs'] in 'Linux':
        parser = argparse.ArgumentParser()

        for i, argv in enumerate(sys.argv[1:]):
            if not argv.__contains__('--'): continue
            parser.add_argument(argv)

        inParInfo = vars(parser.parse_args())

    log.info("[CHECK] inParInfo : {}".format(inParInfo))

    for key, val in inParInfo.items():
        if val is None: continue
        # 전역 변수에 할당
        globalVar[key] = val

    # 전역 변수
    for key, val in globalVar.items():
        if env not in 'local' and key.__contains__('Path') and env and not os.path.exists(val):
            os.makedirs(val)

        globalVar[key] = val.replace('\\', '/')

        log.info("[CHECK] {} : {}".format(key, val))

        # self 변수에 할당
        # setattr(self, key, val)

    return globalVar

# ================================================
# 4. 부 프로그램
# ================================================
class DtaProcess(object):

    # ================================================
    # 요구사항
    # ================================================
    # Python을 이용한 국가 입력 데이터를 기준으로 기준 엑셀파일 현행화

    # cd /DATA/OUTPUT/LSH0554

    # zip -r china-2020-2050.zip china-2020-2050/

    # ps -ef | grep LSH0554 | awk '{print $2}' | xargs kill -9

    # ================================================================================================
    # 환경변수 설정
    # ================================================================================================
    global env, contextPath, prjName, serviceName, log, globalVar

    # env = 'local'  # 로컬 : 원도우 환경, 작업환경 (현재 소스 코드 환경 시 .) 설정
    env = 'dev'  # 개발 : 원도우 환경, 작업환경 (사용자 환경 시 contextPath) 설정
    # env = 'oper'  # 운영 : 리눅스 환경, 작업환경 (사용자 환경 시 contextPath) 설정

    if (platform.system() == 'Windows'):
        contextPath = os.getcwd() if env in 'local' else 'E:/04. TalentPlatform/Github/TalentPlatform-Python'
    else:
        contextPath = os.getcwd() if env in 'local' else '/SYSTEMS/PROG/PYTHON/PyCharm'

    prjName = 'test'
    serviceName = 'LSH0554'

    # 4.1. 환경 변수 설정 (로그 설정)
    log = initLog(env, contextPath, prjName)

    # 4.2. 환경 변수 설정 (초기 변수)
    globalVar = initGlobalVar(env, contextPath, prjName)

    # ================================================================================================
    # 4.3. 초기 변수 (Argument, Option) 설정
    # ================================================================================================
    def __init__(self, inParams):

        log.info("[START] __init__ : {}".format("init"))

        try:
            # 초기 전달인자 설정 (파이썬 실행 시)
            # pyhton3 *.py argv1 argv2 argv3 ...
            initArgument(globalVar, inParams)

        except Exception as e:
            log.error("Exception : {}".format(e))
            raise e
        finally:
            log.info("[END] __init__ : {}".format("init"))

    # ================================================================================================
    # 4.4. 비즈니스 로직 수행
    # ================================================================================================
    def exec(self):

        log.info('[START] {}'.format("exec"))

        try:

            if (platform.system() == 'Windows'):
                pass
            else:
                globalVar['inpPath'] = '/DATA/INPUT'
                globalVar['outPath'] = '/DATA/OUTPUT'
                globalVar['figPath'] = '/DATA/FIG'
                globalVar['updPath'] = '/DATA/CSV'

            sysOpt = {
                'metaList': [
                    "2020_TPL_input2015",
                    "2020_input2015",
                    "2030_1.5_SSP1",
                    "2030_1.5_SSP2",
                    "2030_2_SSP1",
                    "2030_2_SSP2",
                    "2030_SSP1_nopolicy",
                    "2030_SSP2_nopolicy",
                    "2050_1.5_SSP1",
                    "2050_1.5_SSP2",
                    "2050_2_SSP1",
                    "2050_2_SSP2",
                    "2050_SSP1_nopolicy",
                    "2050_SSP2_nopolicy"
                ]
                , 'keyTypeMat': {
                    'anhui': 'ANHU',
                    'beijing': 'BEIJ',
                    'chongqing': 'CHON',
                    'fujian': 'FUJI',
                    'gansu': 'GANS',
                    'guangdong': 'GUAD',
                    'guangxi': 'GUAX',
                    'guizhou': 'GUIZ',
                    'hainan': 'HAIN',
                    'hebei': 'HEBE',
                    'heilongjiang': 'HEIL',
                    'henan': 'HENA',
                    'hubei': 'HUBE',
                    'hunan': 'HUNA',
                    'innermongolia': 'NEMO',
                    'jiangsu': 'JINU',
                    'jiangxi': 'JINX',
                    'jilin': 'JILI',
                    'liaoning': 'LIAO',
                    'ningxia': 'NINX',
                    'qinghai': 'QING',
                    'shaanxi': 'SHAA',
                    'shandong': 'SHND',
                    'shanghai': 'SHAN',
                    'shanxi': 'SHNX',
                    'sichuan': 'SICH',
                    'tianjin': 'TIAN',
                    'tibet': 'TIBE',
                    'xinjiang': 'XING',
                    'yunnan': 'YUNN',
                    'zhejiang': 'ZHEJ'
                }
            }

            # ********************************************************************
            # 테스트
            # ********************************************************************
            # metaInfo = '2015_TPL_input2015'
            # metaInfo = '2019_TPL_input2015'
            # for metaInfo in sysOpt['metaList']:
            #     log.info(f'[CHECK] metaInfo : {metaInfo}')
            #
            #     inpFilePatrn = f'china-20240428/**/ChinaPower{metaInfo}/**/*.xlsx'
            #     inpFile = '{}/{}/{}'.format(globalVar['inpPath'], serviceName, inpFilePatrn)
            #     fileList = sorted(glob.glob(inpFile, recursive=True))
            #
            #     dataL1 = pd.DataFrame()
            #     for fileInfo in fileList:
            #         fileName = os.path.basename(fileInfo)
            #         fileNameSepList = re.split(r'[厂_\.]', fileName)
            #         if fileNameSepList is None or len(fileNameSepList) < 1: continue
            #
            #         keyType = 'None'
            #         try:
            #             keyType = fileNameSepList[5]
            #         except Exception:
            #             pass
            #
            #         dict = {
            #             'keyType': [keyType]
            #             , 'keyTypeAbbr': [keyType[:4].upper()]
            #             , 'fileName': [fileName]
            #         }
            #
            #         dataL1 = pd.concat([dataL1, pd.DataFrame.from_dict(dict)], ignore_index=True)

            # ********************************************************************
            # 국가 입력 데이터
            # ********************************************************************
            # metaInfo = '2030_SSP1_nopolicy'
            # for metaInfo in sysOpt['metaList']:
            for metaInfo in sysOpt['metaList']:

                log.info(f'[CHECK] metaInfo : {metaInfo}')

                # inpFilePatrn = f'china-20240502/**/{metaInfo}/**/*.csv'
                # inpFilePatrn = f'china/**/{metaInfo}/**/*.csv'
                # inpFilePatrn = f'china-20240502/**/{metaInfo}/**/*.csv'
                # inpFilePatrn = f'china-20240509/**/{metaInfo}/**/*.csv'
                inpFilePatrn = f'china-20240528/**/{metaInfo}/**/*.csv'
                inpFile = '{}/{}/{}'.format(globalVar['inpPath'], serviceName, inpFilePatrn)
                fileList = sorted(glob.glob(inpFile, recursive=True))

                # fileInfo = fileList[0]
                dataL1 = pd.DataFrame()
                for fileInfo in fileList:
                    if re.match(r'.*/~\$.*', fileInfo): continue
                    log.info(f'[CHECK] fileInfo : {fileInfo}')

                    fileName = os.path.basename(fileInfo)
                    fileNameSepList = re.split(r'[厂_\.]', fileName)
                    if fileNameSepList is None or len(fileNameSepList) < 1: continue

                    keyType = 'None'
                    try:
                        keyType = fileNameSepList[len(fileNameSepList) - 2]
                    except Exception:
                        pass

                    filePathSepList = re.split(r'[/]', fileInfo)

                    # keyInfo = fileInfo.split('/')[7]
                    keyInfo = filePathSepList[6]
                    if keyInfo is None: continue

                    # if re.search('nopolicy', keyInfo, re.IGNORECASE): continue

                    keyInfoSepList = re.split(r'[厂_]', keyInfo)
                    if keyInfoSepList is None or len(keyInfoSepList) < 1: continue

                    keyYear = 'None'
                    keyIdx = 'None'
                    keyVer = 'None'

                    try:
                        keyYear = keyInfoSepList[0]
                    except Exception:
                        pass

                    try:
                        keyIdx = keyInfoSepList[1]
                    except Exception:
                        pass

                    try:
                        keyVer = keyInfoSepList[2]
                    except Exception:
                        pass

                    data = pd.read_csv(fileInfo)
                    data['keyYear'] = keyYear
                    data['keyIdx'] = keyIdx
                    data['keyVer'] = keyVer
                    data['keyType'] = sysOpt['keyTypeMat'][keyType.lower()]
                    data['fileName'] = fileName

                    dataL1 = pd.concat([dataL1, pd.DataFrame.from_dict(data)], ignore_index=True)

                # dataL1['key'] = dataL1[['keyYear', 'keyIdx', 'keyVer', 'keyType']].apply(lambda row: '_'.join(str(x) for x in row if x is not None), axis=1)

                keyYearList = sorted(set(dataL1['keyYear']))
                keyTypeList = sorted(set(dataL1['keyType']))

                for keyYearInfo in keyYearList:
                    for keyTypeInfo in keyTypeList:

                        # keyYearInfo = '2030'
                        # keyTypeInfo = 'ANHU'

                        log.info(f'[CHECK] keyYearInfo : {keyYearInfo} / keyTypeInfo : {keyTypeInfo}')

                        dataL2 = dataL1.loc[(dataL1['keyYear'] == keyYearInfo) & (dataL1['keyType'] == keyTypeInfo)]
                        if dataL2 is None or len(dataL2) < 1: continue

                        # ********************************************************************
                        # 기준 엑셀파일 읽기
                        # ********************************************************************
                        # inpFilePatrn = f'China/ChinaPower{metaInfo}_*/**/*{keyYearInfo}*_{keyTypeInfo}_*.xlsx'
                        inpFilePatrn = f'china-20240509/**/*_{keyTypeInfo}_*.xlsx'
                        inpFile = '{}/{}/{}'.format(globalVar['inpPath'], serviceName, inpFilePatrn)
                        fileList = sorted(glob.glob(inpFile, recursive=True))
                        if len(fileList) < 1: continue

                        # fileInfo = fileList[0]
                        for fileInfo in fileList:
                            if re.match(r'.*/~\$.*', fileInfo): continue
                            log.info(f'[CHECK] fileInfo : {fileInfo}')

                            # 엑셀 파일 읽기
                            # data_only=True 수식 제거
                            # data_only=False 수식 유지
                            wb = load_workbook(fileInfo, data_only=False)

                            # En_ppl 시트
                            ws = wb['En_ppl']
                            wbData = pd.read_excel(fileInfo, sheet_name='En_ppl', engine='openpyxl', skiprows=2)
                            colNameItem = {cell.value: cell.column_letter for cell in ws[3]}

                            # 초기화
                            for idx, item in wbData.iterrows():
                                for colName, colCell in colNameItem.items():
                                    if colName is None: continue
                                    if re.search('year', colName, re.IGNORECASE): continue
                                    if re.search('Act_abb', colName, re.IGNORECASE): continue
                                    if re.search('None', colName, re.IGNORECASE): continue
                                    if re.search('PP_TOTAL', colName, re.IGNORECASE): continue

                                    cell = ws[f'{colNameItem[colName]}{idx + 4}']
                                    cellFill = cell.fill.start_color.index
                                    if cellFill != '00000000': continue
                                    cell.value = 0.0

                            # keyGrpYearInfo = '2015' if re.search('2015|2016|2017|2018|2019', keyYearInfo) else '2020' if re.search('2020', keyYear) else 'None'
                            # keyGrpYearInfo = keyYearInfo
                            keyGrpYearInfo = '2015' if re.search('2020', keyYearInfo) else keyYearInfo

                            # engType = 'HC3'
                            for idx, item in dataL2.iterrows():
                                # engType = item['EnergyType']
                                engType = item.get('EnergyType')
                                if pd.isna(engType): continue
                                wbDataL1 = wbData.loc[(wbData['year'] == int(keyGrpYearInfo)) & (wbData['Act_abb'] == engType)]
                                if wbDataL1.size < 1: continue
                                rowIdx = wbDataL1.index[0] + 4

                                for colName, colCell in colNameItem.items():
                                    if colName is None: continue
                                    if re.search('year', colName, re.IGNORECASE): continue
                                    if re.search('Act_abb', colName, re.IGNORECASE): continue
                                    if re.search('None', colName, re.IGNORECASE): continue
                                    if re.search('PP_TOTAL', colName, re.IGNORECASE): continue

                                    # 셀 배경 채우기
                                    cell = ws[f'{colNameItem[colName]}{rowIdx}']
                                    cellFill = cell.fill.start_color.index

                                    if cellFill != '00000000': continue
                                    # cell.value = 0.0

                                    colVal = cell.value
                                    selVal = item.get(colName)
                                    if selVal is None: continue
                                    if colVal == selVal: continue

                                    cell.value = selVal
                                    # log.info(f'[CHECK] engType : {engType} / colName : {colName} / colVal : {colVal} / selVal : {selVal} / cell.value : {cell.value}')

                            # fileName = os.path.basename(fileInfo)
                            # srtIdx = fileInfo.index('ChinaPower')
                            srtIdx = fileInfo.index('actPathEneMob')
                            # fileName = fileInfo[srtIdx:]
                            metaInfo2 = re.sub(r'\.', '', metaInfo)
                            metaInfo3 = f'ChinaPower{metaInfo2}'
                            fileName = fileInfo[srtIdx:].replace('ChinaPower2015', metaInfo3)

                            # Main 시트
                            wsMain = wb['Main']
                            # eleList = [keyYearInfo, sKeyIdxInfo, sKeyVerInfo]
                            # eleList = [metaInfo]
                            # eleExtList = [element for element in eleList if element != '']
                            # eleExtMrg = '_'.join(eleExtList)
                            # wsMain[f'B7'].value = f'ChinaPower{eleExtMrg}'
                            wsMain[f'B7'].value = metaInfo3

                            # 기울림체 삭제
                            wsMain[f'B3'].value = metaInfo3
                            orgFont = wsMain[f'B3'].font
                            newFont = copy(orgFont)
                            newFont.italic = False
                            wsMain[f'B3'].font = newFont

                            # saveFile = '{}/{}/{}/{}/{}'.format(globalVar['outPath'], serviceName, 'china-2030-2050', metaInfo3, fileName)
                            saveFile = '{}/{}/{}/{}/{}'.format(globalVar['outPath'], serviceName, 'china-2020-2050', metaInfo3, fileName)
                            os.makedirs(os.path.dirname(saveFile), exist_ok=True)
                            wb.save(saveFile)
                            log.info('[CHECK] saveFile : {}'.format(saveFile))

        except Exception as e:
            log.error("Exception : {}".format(e))
            raise e

        finally:
            log.info('[END] {}'.format("exec"))

# ================================================
# 3. 주 프로그램
# ================================================
if __name__ == '__main__':

    print('[START] {}'.format("main"))

    try:

        # 파이썬 실행 시 전달인자를 초기 환경변수 설정
        inParams = {}

        print("[CHECK] inParams : {}".format(inParams))

        # 부 프로그램 호출
        subDtaProcess = DtaProcess(inParams)

        subDtaProcess.exec()

    except Exception as e:
        print(traceback.format_exc())
        sys.exit(1)

    finally:
        print('[END] {}'.format("main"))
